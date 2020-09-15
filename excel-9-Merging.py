from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()
mysheet = wb.active

# merge cells
mysheet.merge_cells('B2:D3')
mysheet['A1'] = 'cells merged together.'
mysheet.merge_cells('F6:F7')
mysheet['G5'] = 'Two merged cells.'

wb.save('excel-9-Merging.xlsx')
