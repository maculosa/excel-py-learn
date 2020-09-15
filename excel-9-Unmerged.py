import openpyxl

wb = openpyxl.load_workbook('excel-9-Merging.xlsx')
mysheet = wb.active

# ynmerge cells
mysheet.unmerge_cells('B2:D3')
mysheet.unmerge_cells('F6:F7')

wb.save('excel-9-unmerged.xlsx')
