from openpyxl import Workbook
wb = Workbook()     # 一个 workbook 总是至少创建一个工作表

#

ws = wb.active      # 使用默认（第一个）工作表

# 使用 Workbook.create_sheet() 方法创建新的工作表
ws1 = wb.create_sheet('MySheet') # insert at the end (default)
wx2 = wb.create_sheet('MySheet', 0) # insert at first position
wx3 = wb.create_sheet('MySheet', -1) # insert at the penultimate position


ws.title = "New Title"
ws.sheet_properties.tabColor = '1072BA'

ws3 = wb['New Title']

print(wb.sheetnames)
# ['Sheet2', 'New Title', 'Sheet1']

for sheet in wb:
    print(sheet.title)

source = wb.active
target = wb.copy_worksheet(source)