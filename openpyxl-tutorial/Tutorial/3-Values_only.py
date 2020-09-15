import openpyxl

wb = openpyxl.Workbook()
ws = wb.active

# Values only
print('\n# Values only\n')
for row in ws.values:
    for value in row:
        print(value)

# # Both Worksheet.iter_rows() and Worksheet.iter_cols() can take the values_only parameter to return just the cell’s value:
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
   print(row)


# Saving to a file
wb.save('3-Values_only.xlsx')
