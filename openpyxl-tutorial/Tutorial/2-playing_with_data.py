import openpyxl

wb = openpyxl.Workbook()
ws = wb.active

# Accessing one cell
c = ws['A4']
ws['A4'] = 4

d = ws.cell(row=4, column=2, value=10)

# Accessing many cells

cell_range = ws['A1': 'C2']     # Ranges of cells can be accessed using slicing:
## Ranges of rows or columns can be obtained similarly
colC = ws['C']
col_range = ws['C:D']
row10 = ws[10]
row_range = ws[5:10]

## You can also use the Worksheet.iter_rows() method:
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print('row: ', cell)

## Likewise the Worksheet.iter_cols() method will return columns:
for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
    for cell in col:
        print('col: ', cell)


ws = wb.active
ws['C9'] = 'hello world'
tu_rows = tuple(ws.rows)
print(tu_rows)

tu_cols = tuple(ws.columns)
print(tu_cols)

# Values only
print('\n# Values only\n')
for row in ws.values:
    for value in row:
        print(value)

# # Both Worksheet.iter_rows() and Worksheet.iter_cols() can take the values_only parameter to return just the cell’s value:
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
   print(row)

# Data storage
print('\n# Data storage\n')

c.value = 'hello, world'
print(c.value)

d.value = 3.14
print(d.value)

# Saving to a file
wb.save('2-play_with_data.xlsx')
