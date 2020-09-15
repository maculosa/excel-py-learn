import openpyxl

wb = openpyxl.Workbook()
ws = wb.active

# Accessing one cell
c = ws['A4']
ws['A4'] = 4

d = ws.cell(row=4, column=2, value=10)

# Data storage
print('\n# Data storage\n')

c.value = 'hello, world'
print(c.value)

d.value = 3.14
print(d.value)

# Saving to a file
wb.save('4-Data_storage.xlsx')
