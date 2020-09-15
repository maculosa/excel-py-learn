import openpyxl

mywb = openpyxl.Workbook()
print(mywb.get_sheet_names())

# 创建 sheet
mywb.create_sheet()

print(mywb.get_sheet_names())

mywb.create_sheet(index=0, title='1st Sheet')
print(mywb.get_sheet_names())

mywb.create_sheet(index=2, title='2nd Sheet')
print(mywb.get_sheet_names())
